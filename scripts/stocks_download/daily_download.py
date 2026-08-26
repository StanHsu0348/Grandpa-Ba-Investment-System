"""
每日自動下載美股/台股清單，覆蓋 data/uslist.xlsx、data/twlist.xlsx，
並把當天下載到的資料另存一份到 data/history/ 留存歷史記錄。

複用 setup_login.py 存下的登入狀態 (auth_state.json)，不需要每次重新登入。
如果 session 已經過期（被導回登入頁），會印出訊息並以非 0 狀態碼結束，
這時需要重新執行 setup_login.py 手動登入一次。

下載成功後，會自動 git add/commit/push 到 GitHub（private repo），
讓部署在 Streamlit Community Cloud 上的版本跟著自動重新部署、拿到最新資料。
git push 失敗（例如沒網路、repo 尚未設定 remote）不會讓整支腳本失敗，
只會記錄在 log 裡，下載到的資料仍然會留在本機 data/ 目錄。

用法:
    python daily_download.py
"""
import datetime
import shutil
import subprocess
import sys
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

BASE_DIR = Path(__file__).resolve().parent
REPO_DIR = BASE_DIR.parent.parent
STATE_FILE = BASE_DIR / "auth_state.json"
LOG_FILE = BASE_DIR / "download_log.txt"
DATA_DIR = REPO_DIR / "data"
HISTORY_DIR = DATA_DIR / "history"
DOWNLOAD_URL = "https://stocks.ddns.net/App/DownloadList.aspx"

# (下載頁面上的按鈕 id, 存檔檔名, 驗證用的工作表名稱)
TARGETS = [
    ("ctl00_ContentPlaceHolder1_Export", "uslist.xlsx", "US"),       # 美股
    ("ctl00_ContentPlaceHolder1_Linkbutton1", "twlist.xlsx", "TW"),  # 台股
]


def log(msg: str) -> None:
    line = f"[{datetime.datetime.now():%Y-%m-%d %H:%M:%S}] {msg}"
    print(line)
    with LOG_FILE.open("a", encoding="utf-8") as f:
        f.write(line + "\n")


def verify_xlsx(path: Path, expected_sheet: str) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ok = expected_sheet in wb.sheetnames and wb[expected_sheet].max_row > 1
        wb.close()
        return ok
    except Exception as e:
        log(f"  驗證失敗: {e}")
        return False


def git_sync(today_str: str) -> None:
    """把更新後的 data/ 推到 GitHub，讓 Streamlit Cloud 自動重新部署。
    push 失敗（沒網路、remote 未設定等）只記錄 log，不影響本機資料已更新的事實。
    """
    try:
        subprocess.run(
            ["git", "add", "data"], cwd=REPO_DIR, check=True, capture_output=True, text=True,
        )
        diff = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=REPO_DIR)
        if diff.returncode == 0:
            log("  資料內容與上次相同，無需 commit。")
            return

        subprocess.run(
            ["git", "commit", "-m", f"自動更新股票資料 {today_str}"],
            cwd=REPO_DIR, check=True, capture_output=True, text=True,
        )
        push = subprocess.run(["git", "push"], cwd=REPO_DIR, capture_output=True, text=True)
        if push.returncode == 0:
            log("  已推送到 GitHub，Streamlit Cloud 會自動重新部署。")
        else:
            log(f"  git push 失敗（資料仍留在本機，不影響今天使用）: {push.stderr.strip()}")
    except subprocess.CalledProcessError as e:
        detail = e.stderr.strip() if e.stderr else str(e)
        log(f"  git 同步失敗: {detail}")
    except FileNotFoundError:
        log("  找不到 git 指令，略過自動同步。")


def main() -> int:
    if not STATE_FILE.exists():
        log("找不到 auth_state.json，請先執行 setup_login.py 手動登入一次。")
        return 1

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    today_str = datetime.date.today().strftime("%Y%m%d")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(storage_state=str(STATE_FILE))
        page = context.new_page()
        page.goto(DOWNLOAD_URL, wait_until="networkidle")

        if "Login.aspx" in page.url:
            log("Session 已過期（被導回登入頁），請重新執行 setup_login.py 手動登入一次。")
            browser.close()
            return 1

        all_ok = True
        any_updated = False
        for button_id, filename, sheet_name in TARGETS:
            log(f"下載中: {filename} ...")
            tmp_path = DATA_DIR / f".tmp_{filename}"
            try:
                with page.expect_download(timeout=60000) as dl_info:
                    page.click(f"#{button_id}")
                download = dl_info.value
                download.save_as(str(tmp_path))
            except Exception as e:
                log(f"  下載失敗: {e}")
                all_ok = False
                continue

            if not verify_xlsx(tmp_path, sheet_name):
                log(f"  檔案驗證失敗（可能下載到錯誤內容），保留原本的 {filename} 不覆蓋。")
                tmp_path.unlink(missing_ok=True)
                all_ok = False
                continue

            final_path = DATA_DIR / filename
            tmp_path.replace(final_path)

            stem = Path(filename).stem
            history_path = HISTORY_DIR / f"{stem}_{today_str}.xlsx"
            shutil.copy2(final_path, history_path)

            log(f"  完成，已更新 {final_path}（歷史備份: {history_path}）")
            any_updated = True

        browser.close()

    if any_updated:
        git_sync(today_str)

    return 0 if all_ok else 1


if __name__ == "__main__":
    sys.exit(main())
